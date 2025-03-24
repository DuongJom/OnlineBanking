from flask import session
import io
import pandas as pd
from openpyxl.styles import Alignment, PatternFill
from openpyxl import load_workbook

from enums.card_type import CardType
from enums.mime_type import MIMEType
from enums.collection import CollectionType
from enums.role_type import RoleType
from enums.data_type import DataType
from message import messages_success, messages_failure
from models.database import Database
from models.card import Card
from models.user import User
from models.account import Account
from helpers.helpers import issue_new_card, generate_login_info, getMIMETypeValue, get_max_id
from filters import format_id

db = Database().get_db()
accounts = db[CollectionType.ACCOUNTS.value]
users = db[CollectionType.USERS.value]
cards = db[CollectionType.CARDS.value]
branches = db[CollectionType.BRANCHES.value]
employees = db[CollectionType.EMPLOYEES.value]
news = db[CollectionType.NEWS.value]

lst_formated_column = {
    CollectionType.ACCOUNTS.value: ['ID', 'Account Owner', 'Branch ID','Balance']
}

def create_accounts(data):
    admin_id = session.get("account_id")
    total = len(data)
    counter = 0
    row_index = 1
    response_data = {"total": total, "logs": []}
    res = None

    for accData in data:
        fullname = accData["Fullname"]
        gender = int(accData["Gender"].split('.')[0])
        phone = str(accData["Phone"]).strip()
        email = accData["Email"].strip()
        branch = int(accData["Branch"].split('.')[0])
        address = accData["Address"]
        card_info = issue_new_card()
        login_info = generate_login_info(email, phone)     

        is_email_exits = True if users.find_one({"Email": email}) else False
        is_phone_exits = True if users.find_one({"Phone": phone}) else False
        is_error = False
        res = {
            'status': "success", 
            'message': messages_success["create_success"].format(login_info['Username']), 
            "row_index": row_index
        }

        if is_email_exits:
            is_error = True
            res = {
                'status': "error", 
                'message': messages_failure["email_existed"].format(email), 
                "row_index": row_index
            }
        if is_phone_exits:
            is_error = True
            res = {
                'status': "error", 
                'message': messages_failure["phone_existed"].format(phone), 
                "row_index": row_index
            }

        if not is_error:
            new_card_id = get_max_id(database=db, collection_name=CollectionType.CARDS.value)
            new_card = Card(
                id=new_card_id, 
                cardNumber=card_info['cardNumber'], 
                cvv=card_info['cvvNumber'], 
                type=CardType.CREDITS.value,
                createdBy = admin_id
            )

            new_user_id = get_max_id(database=db, collection_name=CollectionType.USERS.value)
            new_user = User(
                id=new_user_id, 
                name=fullname, 
                sex=int(gender), 
                address=address.strip(), 
                phone=phone, 
                email=email, 
                card=[new_card_id,],
                createdBy = admin_id
            )

            new_account_id = get_max_id(database=db, collection_name=CollectionType.ACCOUNTS.value)
            new_account = Account(
                id=new_account_id,
                accountNumber=card_info['accountNumber'], 
                branch=int(branch), 
                user=new_user_id, 
                username=login_info['Username'], 
                password=login_info['Password'], 
                role=RoleType.USER.value, 
                transferMethod=[1], 
                loginMethod=[1],
                createdBy = admin_id
            )

            cards.insert_one(new_card.to_json())
            users.insert_one(new_user.to_json())
            accounts.insert_one(new_account.to_json())

        row_index += 1
        response_data["logs"].append(res)
        response_data.update({"success_create": counter})
    return response_data

def generate_export_account(file_type, criteria, data_type):
    if data_type == DataType.ACCOUNT.value:
        collection_name = CollectionType.ACCOUNTS.value
        items = get_export_account(criteria)
    df = pd.json_normalize(items)
    output = io.BytesIO()
    mime = None
    file_type = int(file_type)

    if file_type == MIMEType.CSV.value:
        df.to_csv(output, index=False, encoding='utf-8')
        output.seek(0) 
        mime = getMIMETypeValue(MIMEType.CSV)

    elif file_type == MIMEType.JSON.value:
        df = df.applymap(lambda x: x if isinstance(x, str) else str(x))
        df = df.applymap(lambda x: x.encode('utf-8', errors='replace').decode('utf-8'))
        df.to_json(output, index=False, force_ascii=False, orient='records', lines=True)
        output.seek(0)
        mime = getMIMETypeValue(MIMEType.JSON)

    elif file_type == MIMEType.EXCEL_XLSX.value:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
        output.seek(0)  
        temp_output = io.BytesIO()
        wb = load_workbook(output)
        ws = wb.active

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for col in ws.columns:
            col_name = col[0].value 
            if col_name in lst_formated_column[collection_name]:
                if pd.api.types.is_integer_dtype(df[col_name]): 
                    for cell in col[1:]: 
                        cell.value = format_id(cell.value, 5)
                        cell.alignment = Alignment(horizontal='right')

                if pd.api.types.is_float_dtype(df[col_name]): 
                    for cell in col[1:]: 
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')

            max_length = max(len(str(cell.value)) for cell in col if cell.value) + 5
            ws.column_dimensions[col[0].column_letter].width = max_length

        wb.save(temp_output)
        temp_output.seek(0)
        mime = getMIMETypeValue(MIMEType.EXCEL_XLSX)

        return {'mime': mime, 'output': temp_output}  

    return {'mime': mime, 'output': output}

def get_export_account(criteria):
    pipeline = [
        { "$match": criteria},
        { 
            "$lookup": {
                "from": "login_methods",
                "localField": "LoginMethod",
                "foreignField": "_id",
                "as": "login_method_docs"
            }
        },
        {
            "$lookup": {
                "from": "transfer_methods",
                "localField": "TransferMethod",
                "foreignField": "_id",
                "as": "transfer_method_docs"
            }
        },
        {
            "$lookup": {
                "from": "users",
                "localField": "AccountOwner",
                "foreignField": "_id",
                "as": "owner_doc"
            }
        },
        {
            "$lookup": {
                "from": "branches",
                "localField": "Branch",
                "foreignField": "_id",
                "as": "branch_doc"
            }
        },
        {
            "$lookup": {
                "from": "roles",
                "localField": "Role",
                "foreignField": "_id",
                "as": "role_doc"
            }
        },
        
        {
            "$project": {
                "_id": 0,
                "ID": "$_id",
                "Account Number": "$AccountNumber",
                "Account Owner": "$AccountOwner",
                "Owner Name": {
                    "$ifNull": [{"$arrayElemAt": ["$owner_doc.Name", 0]}, None]
                },
                "Branch ID": "$Branch",
                "Branch Name": {
                    "$ifNull": [{"$arrayElemAt": ["$branch_doc.BranchName", 0]}, None]
                },
                "Balance": "$Balance",
                "Currency": {
                    "$cond": {
                        "if": {"$eq": ["$Currency", 0]},
                        "then": "VND",
                        "else": {
                            "$cond": {
                                "if": {"$eq": ["$IsDeleted", 1]},
                                "then": "USD",
                                "else": "EUR"
                            }
                        }
                    }
                },
                "Deleted Type": {
                    "$cond": {
                        "if": {"$eq": ["$IsDeleted", 0]},
                        "then": "Available",
                        "else": {
                            "$cond": {
                                "if": {"$eq": ["$IsDeleted", 9]},
                                "then": "Deleted",
                                "else": "Unknown"
                            }
                        }
                    }
                },
                "Login Methods": {
                    "$ifNull": [{
                        "$map": {
                            "input": "$login_method_docs",
                            "as": "method",
                            "in": "$$method.MethodName"
                        }
                    }, []]
                },
                "Role": {
                    "$ifNull": [{"$arrayElemAt": ["$role_doc.RoleName", 0]}, None]
                },  
                "Transfer Methods": {
                    "$ifNull": [{
                        "$map": {
                            "input": "$transfer_method_docs",
                            "as": "method",
                            "in": "$$method.MethodName"
                        }
                    }, []]
                },
                "Username": "$Username",
            }
        }
    ]            

    lst_account = lst_account = list(accounts.aggregate(pipeline))
    for acc in lst_account:
        acc['Login Methods'] = ', '.join(acc['Login Methods'])
        acc['Transfer Methods'] = ', '.join(acc['Transfer Methods'])
    
    return lst_account


def get_account(id):
    pipeline = [
        {
            "$match": {
                "_id": id
            }
        },
        {
            "$lookup": {
                "from": "login_methods",
                "localField": "LoginMethod",
                "foreignField": "_id",
                "as": "login_method_docs"
            }
        },
        {
            "$lookup": {
                "from": "transfer_methods",
                "localField": "TransferMethod",
                "foreignField": "_id",
                "as": "transfer_method_docs"
            }
        },
        {
            "$lookup": {
                "from": "users",
                "localField": "AccountOwner",
                "foreignField": "_id",
                "as": "owner_doc"
            }
        },
        {
            "$lookup": {
                "from": "branches",
                "localField": "Branch",
                "foreignField": "_id",
                "as": "branch_doc"
            }
        },
        {
            "$lookup": {
                "from": "roles",
                "localField": "Role",
                "foreignField": "_id",
                "as": "role_doc"
            }
        },
        {
            "$project": {
                "_id": 1,
                "Username": 1,
                "AccountNumber": 1,
                "Balance": 1,
                "IsDeleted": 1,
                "Owner_name": {
                    "$ifNull": [{"$arrayElemAt": ["$owner_doc.Name", 0]}, None]
                },
                "Branch_name": {
                    "$ifNull": [{"$arrayElemAt": ["$branch_doc.BranchName", 0]}, None]
                },
                "Role_name": {
                    "$ifNull": [{"$arrayElemAt": ["$role_doc.RoleName", 0]}, None]
                },
                "lst_login_method": {
                    "$ifNull": [{
                        "$map": {
                            "input": "$login_method_docs",
                            "as": "method",
                            "in": "$$method.MethodName"
                        }
                    }, []]
                },
                "lst_transfer_method": {
                    "$ifNull": [{
                        "$map": {
                            "input": "$transfer_method_docs",
                            "as": "method",
                            "in": "$$method.MethodName"
                        }
                    }, []]
                }
            }
        }
    ]
    account = dict(next(accounts.aggregate(pipeline)))
    return account